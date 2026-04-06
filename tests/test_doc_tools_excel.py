# Copyright (c) 2025 Kenneth Stott
# Canary: 6a396d74-c4eb-4d41-8169-9268f9f268c6
#
# This source code is licensed under the Business Source License 1.1
# found in the LICENSE file in the root directory of this source tree.
#
# NOTICE: Use of this software for training artificial intelligence or
# machine learning models is strictly prohibited without explicit written
# permission from the copyright holder.

"""Tests for Excel/XLSX document loading in doc_tools module."""

from __future__ import annotations

import io
import pytest
from pathlib import Path
from unittest.mock import Mock, patch

from constat.core.config import Config
from constat.discovery.doc_tools import DocumentDiscoveryTools
from constat.discovery.doc_tools import _transport as _doc_transport
from constat.discovery.doc_tools._file_extractors import (
    _extract_xlsx_text,
    _extract_xlsx_text_from_bytes,
)


# Reset module-level HTTP session singleton between tests
@pytest.fixture(autouse=True)
def _reset_http_session():
    _doc_transport._http_session = None
    yield
    _doc_transport._http_session = None


# =============================================================================
# Helper Functions for Creating Test Excel Files
# =============================================================================


def create_minimal_xlsx() -> bytes:
    """
    Create a minimal valid Excel spreadsheet.

    Returns:
        XLSX file bytes
    """
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Sample data"

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


def create_xlsx_with_content(
    sheets: dict[str, list[list]] = None,  # {sheet_name: [[row1], [row2], ...]}
) -> bytes:
    """
    Create an Excel spreadsheet with specified content.

    Args:
        sheets: Dict mapping sheet names to row data (list of lists)

    Returns:
        XLSX file bytes
    """
    from openpyxl import Workbook
    from io import BytesIO

    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active

    if sheets:
        first = True
        for sheet_name, rows in sheets.items():
            if first:
                ws = default_sheet
                ws.title = sheet_name
                first = False
            else:
                ws = wb.create_sheet(title=sheet_name)

            for row_idx, row in enumerate(rows, 1):
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()


# =============================================================================
# Excel Spreadsheet (XLSX) Tests
# =============================================================================


class TestXlsxTextExtraction:
    """Tests for _extract_xlsx_text and _extract_xlsx_text_from_bytes methods."""

    def test_extract_basic_content(self):
        """Test extracting basic content from Excel spreadsheet."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Sheet1": [["Hello", "World"], ["Data", "Here"]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "Hello | World" in result
        assert "Data | Here" in result

    def test_extract_with_sheet_markers(self):
        """Test that sheets have [Sheet: name] markers."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"MySheet": [["Content"]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "[Sheet: MySheet]" in result

    def test_extract_multi_sheet(self):
        """Test extracting content from multiple sheets."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={
                "Sales": [["Product", "Revenue"], ["Widget", "1000"]],
                "Costs": [["Item", "Cost"], ["Materials", "500"]],
            }
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "[Sheet: Sales]" in result
        assert "[Sheet: Costs]" in result
        assert "Product | Revenue" in result
        assert "Item | Cost" in result

    def test_skip_empty_rows(self):
        """Test that empty rows are skipped."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Sheet1": [
                ["Data1", "Data2"],
                [None, None],  # Empty row
                ["Data3", "Data4"],
            ]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "Data1 | Data2" in result
        assert "Data3 | Data4" in result
        # Empty row should not produce extra blank lines

    def test_extract_numeric_values(self):
        """Test extracting numeric values from Excel."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Numbers": [[1, 2, 3], [4.5, 5.5, 6.5]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "1 | 2 | 3" in result
        assert "4.5 | 5.5 | 6.5" in result

    def test_extract_from_file_path(self, tmp_path):
        """Test extracting content from Excel file path."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["File", "Content"]]}
        )
        xlsx_path = tmp_path / "test.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        result = _extract_xlsx_text(xlsx_path)

        assert "File | Content" in result


class TestXlsxLoadingViaFileType:
    """Tests for loading Excel files through type='file' with .xlsx extension."""

    def test_load_xlsx_via_file_type(self, tmp_path):
        """Test that XLSX files are loaded correctly via file type auto-detection."""
        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["Auto-detected", "Excel"]]}
        )
        xlsx_path = tmp_path / "test.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        config = Config(
            documents={
                "test_excel": {
                    "type": "file",
                    "path": str(xlsx_path),
                    "description": "Test Excel spreadsheet",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        result = tools.get_document("test_excel")

        assert "content" in result
        assert "Auto-detected | Excel" in result["content"]
        assert result["format"] == "text"


class TestXlsxLoadingViaXlsxType:
    """Tests for loading Excel files through type='xlsx'."""

    def test_load_xlsx_via_xlsx_type_with_path(self, tmp_path):
        """Test that XLSX files are loaded correctly via xlsx type with path."""
        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["Direct", "Type"]]}
        )
        xlsx_path = tmp_path / "test.xlsx"
        xlsx_path.write_bytes(xlsx_bytes)

        config = Config(
            documents={
                "direct_excel": {
                    "type": "xlsx",
                    "path": str(xlsx_path),
                    "description": "Direct xlsx type document",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        result = tools.get_document("direct_excel")

        assert "content" in result
        assert "Direct | Type" in result["content"]
        assert result["format"] == "text"

    def test_load_xlsx_via_xlsx_type_with_url(self):
        """Test loading Excel file via xlsx type with URL."""
        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["URL", "Content"]]}
        )

        config = Config(
            documents={
                "url_excel": {
                    "type": "xlsx",
                    "url": "https://example.com/spreadsheet.xlsx",
                    "description": "URL-based Excel spreadsheet",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        mock_response = Mock()
        mock_response.content = xlsx_bytes
        mock_response.raise_for_status = Mock()

        with patch('constat.discovery.doc_tools._transport._get_http_session', return_value=Mock(get=Mock(return_value=mock_response))):
            result = tools.get_document("url_excel")

            assert "content" in result
            assert "URL | Content" in result["content"]
            assert result["format"] == "text"

    def test_missing_xlsx_file_raises_error(self):
        """Test that missing Excel file raises error."""
        config = Config(
            documents={
                "missing": {
                    "type": "xlsx",
                    "path": "/nonexistent/file.xlsx",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        result = tools.get_document("missing")

        assert "error" in result


class TestXlsxLoadingViaHttpType:
    """Tests for loading Excel files through type='http' with XLSX content-type."""

    def test_load_xlsx_via_http_content_type(self):
        """Test that XLSX is detected via HTTP content-type header."""
        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["HTTP", "Detection"]]}
        )

        config = Config(
            documents={
                "http_excel": {
                    "type": "http",
                    "url": "https://example.com/spreadsheet.xlsx",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        mock_response = Mock()
        mock_response.content = xlsx_bytes
        mock_response.headers = {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
        mock_response.raise_for_status = Mock()

        with patch('constat.discovery.doc_tools._transport._get_http_session', return_value=Mock(get=Mock(return_value=mock_response))):
            result = tools.get_document("http_excel")

            assert "content" in result
            assert "HTTP | Detection" in result["content"]

    def test_load_xlsx_via_http_url_extension(self):
        """Test that XLSX is detected via .xlsx URL extension."""
        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["URL", "Extension"]]}
        )

        config = Config(
            documents={
                "url_excel": {
                    "type": "http",
                    "url": "https://example.com/spreadsheet.xlsx",
                }
            }
        )
        tools = DocumentDiscoveryTools(config)

        mock_response = Mock()
        mock_response.content = xlsx_bytes
        mock_response.headers = {"content-type": "application/octet-stream"}
        mock_response.raise_for_status = Mock()

        with patch('constat.discovery.doc_tools._transport._get_http_session', return_value=Mock(get=Mock(return_value=mock_response))):
            result = tools.get_document("url_excel")

            assert "content" in result
            assert "URL | Extension" in result["content"]


class TestXlsxEdgeCases:
    """Edge case tests for Excel spreadsheet handling."""

    def test_empty_xlsx_spreadsheet(self):
        """Test handling of empty Excel spreadsheet."""
        from openpyxl import Workbook
        from io import BytesIO

        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        wb = Workbook()
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        xlsx_bytes = output.read()

        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        # Empty spreadsheet should return empty string
        assert isinstance(result, str)

    def test_xlsx_with_special_characters(self):
        """Test Excel spreadsheet with special characters."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Special": [["< > & \" '", "\u00e9\u00e8\u00f1"]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "< > & \"" in result

    def test_xlsx_with_none_values(self):
        """Test Excel spreadsheet with None values."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Data": [["Value", None, "Another"]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        # None values should be represented as empty strings
        assert "Value" in result
        assert "Another" in result

    def test_xlsx_single_cell(self):
        """Test Excel spreadsheet with single cell."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        xlsx_bytes = create_xlsx_with_content(
            sheets={"Single": [["Only cell"]]}
        )
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        assert "Only cell" in result

    def test_xlsx_many_sheets(self):
        """Test Excel spreadsheet with many sheets."""
        config = Config(documents={})
        tools = DocumentDiscoveryTools(config)

        sheets = {f"Sheet{i}": [[f"Content {i}"]] for i in range(1, 11)}
        xlsx_bytes = create_xlsx_with_content(sheets=sheets)
        result = _extract_xlsx_text_from_bytes(xlsx_bytes)

        for i in range(1, 11):
            assert f"[Sheet: Sheet{i}]" in result
            assert f"Content {i}" in result
